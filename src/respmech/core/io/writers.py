"""Write batch results to disk (Excel data + processed CSV + summary + figures).

Consumers of the core's results — the core itself writes nothing. Output layout
matches the legacy tool: ``<out>/data/Average breathdata.xlsx``,
``<out>/data/<file>.breathdata.xlsx`` and ``<out>/data/<file> – Processed data.csv``.

On top of that the writer adds the pieces a *publishable* analysis needs, all built
from the finished result at write time so the golden-pinned result tables are never
touched:

* each workbook carries a **Units** sheet (P10/P21) and a **Provenance** sheet
  (P21) alongside its Data sheet;
* the breath workbook carries a **normalised-EMG** sheet when normalisation is on
  (P14);
* a **Cohort summary** workbook reports mean ± SD / n / CV% across files, and by
  group (P8/P15);
* **diagnostic figures** are drawn into ``<out>/diagnostics/`` (P11);
* two provenance files — ``analysis-used.toml`` and ``run-report.txt`` — are dropped
  in ``<out>/`` so a result is never orphaned from its settings (P7).
"""
from __future__ import annotations

import os
from datetime import datetime

import pandas as pd

from respmech import __version__
from respmech.core import quantities as _units
from respmech.core.summary import build_cohort_summary, normalize_emg_table

_CREATED = f"Created with RespMech v{__version__} (github.com/emilwalsted/respmech)"


def _version_df():
    return pd.DataFrame(
        {"Created": _CREATED, "Website": "https://github.com/emilwalsted/respmech"},
        index=[0]).T.rename(columns={0: "Version info"})


def _wob_mode_text(settings):
    """The work-of-breathing source, in the two words used everywhere this is named
    (D22, UI-overhaul: the Preview & QC table header, the Provenance row and the Units
    Note below all say the SAME thing, deliberately, per the ticket's own wording —
    'let the same qualification follow' — rather than each inventing its own phrasing
    that could drift out of sync on a future edit)."""
    return ("averaged breath" if settings.processing.wob.calc_from == "average"
           else "individual breaths")


def _units_df(columns, settings=None):
    """Unit per column, plus a Note that names the work-of-breathing source on the
    wob* columns (D22, UI-overhaul) — the only columns whose meaning changes with a
    setting the sheet would otherwise say nothing about (a reader opening this sheet
    to check units has no other way to learn that ``wobtotal`` et al. are one
    whole-file value repeated on every breath in the per-file breathdata sheet, or —
    in the cross-file Average breathdata sheet — that every file's own value came from
    the same source). ``settings`` is optional: callers that only need units (there
    are none left in this codebase, kept for the signature's own sake) get an empty
    Note column instead of a crash."""
    um = _units.units_map(columns)
    note = {}
    if settings is not None:
        wob_desc = _wob_mode_text(settings)
        for c in um:
            if str(c).lower().startswith("wob"):
                note[c] = f"Work of breathing from: {wob_desc}"
    return pd.DataFrame({"Column": list(um.keys()), "Unit": list(um.values()),
                         "Note": [note.get(c, "") for c in um.keys()]})


def _provenance_rows(settings, when):
    ts = (when or datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
    ip = settings.input
    rows = [("RespMech version", __version__),
            ("Generated", ts),
            ("Input folder", ip.folder),
            ("Input pattern", ip.files),
            ("Sampling frequency (Hz)", ip.format.sampling_frequency),
            ("Breath separation", f"{settings.processing.segmentation.method}, "
                                  f"buffer {settings.processing.segmentation.buffer}"),
            # D22 (UI-overhaul): the same "average vs individual" choice that makes the
            # Preview & QC table's wob* columns either one repeated value or real
            # per-breath variation — named here so it survives into the written file,
            # not only on screen.
            ("Work of breathing", _wob_mode_text(settings)),
            ("Drift correction", settings.processing.volume.correct_drift),
            ("EMG normalisation", settings.processing.emg.normalization)]
    if ip.channels.entropy:
        # D11 (UI-overhaul): same m/r a reader would need for a methods section, in the same
        # words as the Setup screen's own read-out (settings_screen.py's ent_caption) — only
        # added when entropy is actually computed (an empty channel list means it is not).
        ent = settings.processing.entropy
        rows.append(("Sample entropy", f"m = {ent.epochs - 1}, r = {ent.tolerance:g} × SD"))
    rows.append(("Settings snapshot", "analysis-used.toml"))
    return pd.DataFrame(rows, columns=["Key", "Value"])


_WEBSITE = "https://github.com/emilwalsted/respmech"


def _autofit(writer):
    """Cosmetic, golden-safe polish (audit #29): size each column to its widest cell and
    turn the website value into a clickable hyperlink. Only touches presentation — cell
    VALUES are unchanged, so the golden (which reads values via pandas) is unaffected."""
    from openpyxl.utils import get_column_letter
    for ws in writer.book.worksheets:
        widths = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                text = str(cell.value)
                widths[cell.column] = max(widths.get(cell.column, 0), len(text))
                if text == _WEBSITE and not cell.hyperlink:
                    cell.hyperlink = _WEBSITE
                    cell.style = "Hyperlink"
        for col, w in widths.items():
            ws.column_dimensions[get_column_letter(col)].width = min(w + 2, 80)


def _write_xlsx(df: pd.DataFrame, path: str, settings=None, when=None, extra_sheets=None):
    """Write a Data sheet plus Units, any extra sheets, Provenance and Version.

    Only the Data sheet content is load-bearing (the golden suite pins the DataFrame,
    not the workbook); the extra sheets are additive context for a reader."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Data", index=False)
        _units_df(df.columns, settings).to_excel(writer, sheet_name="Units", index=False)
        for name, edf in (extra_sheets or {}).items():
            edf.to_excel(writer, sheet_name=name, index=False)
        if settings is not None:
            _provenance_rows(settings, when).to_excel(writer, sheet_name="Provenance", index=False)
        _version_df().to_excel(writer, sheet_name="Version", index=False)
        _autofit(writer)


def write_batch(result, settings, outputfolder: str, when: datetime | None = None,
                progress=None, cohort_outputs: bool = True) -> list[str]:
    """Write all enabled outputs plus summary, figures and provenance; returns the
    list of files written.

    ``progress`` is an optional callback taking a ``pipeline.ProgressEvent``. The writing
    phase used to be entirely silent — on a batch with diagnostics on, figure writing alone
    is far longer than the compute it follows, so the GUI looked frozen from the last file's
    "done" to "Finished". Emitting ``stage`` events here lets the Run screen keep the log
    scrolling and the busy bar animating. The events reuse the same vocabulary the compute
    phase already emits, so the Run screen renders them through the existing path; ``None``
    (the CLI's default) keeps the old silent behaviour.

    ``cohort_outputs`` (default True, byte-identical to the old behaviour for every existing
    caller — the CLI has no ``only_files`` path at all) gates every output that represents the
    WHOLE study rather than one file: ``data/Average breathdata.xlsx``, ``data/Cohort
    summary.xlsx``, the cohort Campbell figure, and whether ``run-report.txt``/
    ``analysis-used.toml`` are treated as this folder's definitive record. ``BatchWorker``
    sets it False (via ``core.pipeline.is_subset_run``) whenever a run is restricted to fewer
    files than the study actually has, so a "Process & write this file"/"Re-run failed" click
    can never again silently rebuild the study's cohort-level results — and its provenance —
    from a fraction of it (ticket A05). Per-file outputs (this file's breathdata workbook,
    processed CSV, diagnostic figures) are written exactly as they would be in a full run;
    only the cohort-level and provenance files change behaviour."""
    # Resolved ONCE: both _write_manifest and _write_run_report independently fell back to
    # datetime.now() when `when` is None (which every real caller — the CLI, BatchWorker —
    # does), so their (partial, <timestamp>) filenames could disagree by a second if the
    # clock ticked over between the two calls. Resolving here makes every timestamp in this
    # write — including the per-file Provenance sheets, which already took `when` — the same
    # instant, not four-plus independent readings of the clock.
    when = when or datetime.now()

    # Lazy import: keeps ``core.pipeline`` (and scipy/pandas via it) out of GUI startup — by
    # the time write_batch runs, run_batch has already imported it, so this is free. See the
    # Wave-1.4 lazy-import note in ui/workers.py.
    from respmech.core.pipeline import ProgressEvent

    def _emit(message, file=None):
        if progress is not None:
            progress(ProgressEvent("stage", file=file, message=message))

    datadir = os.path.join(outputfolder, "data")
    os.makedirs(datadir, exist_ok=True)
    written = []

    if settings.output.data.save_breath_by_breath:
        _emit("writing breath-by-breath data")
        for fname, fr in result.ok_files.items():
            p = os.path.join(datadir, f"{fname}.breathdata.xlsx")
            extra = {}
            norm = normalize_emg_table(fr.breaths_table, settings)   # P14
            if norm is not None and len(norm):
                extra["EMG normalised"] = norm
            _write_xlsx(fr.breaths_table, p, settings=settings, when=when, extra_sheets=extra)
            written.append(p)

    if settings.output.data.save_processed:
        for fname, fr in result.ok_files.items():
            if fr.processed is not None:
                _emit(f"writing processed data — {fname}", file=fname)
                p = os.path.join(datadir, f"{fname} – Processed data.csv")
                fr.processed.to_csv(p, index=False)
                written.append(p)

    if settings.output.data.save_average and result.average_table is not None and cohort_outputs:
        _emit("writing average + cohort summary")
        p = os.path.join(datadir, "Average breathdata.xlsx")
        _write_xlsx(result.average_table, p, settings=settings, when=when)
        written.append(p)
        written += _write_cohort_summary(result, settings, datadir, when)   # P8/P15

    _emit("writing diagnostic figures (the slow step)")
    # Per-file figure callback — fires only when figures run in-process (a packaged build, or
    # RESPMECH_NO_FIGURE_SUBPROCESS=1). In the child-process path it is simply not passed
    # across, so those runs get the single message above plus the animated busy bar.
    fig_progress = (lambda fname: _emit(f"figures — {fname}", file=fname)) if progress else None
    fig_written, fig_failures = _write_figures(result, settings, outputfolder,
                                               progress=fig_progress,
                                               cohort_outputs=cohort_outputs)  # P11
    written += fig_written

    # provenance — always written for a full run, so a folder of results carries its own
    # recipe (P7). For a subset, the full run's provenance is protected instead (A05).
    manifest_path = _write_manifest(settings, outputfolder, cohort_outputs=cohort_outputs,
                                    when=when)
    if manifest_path:
        written.append(manifest_path)
    manifest_name = os.path.basename(manifest_path) if manifest_path else "analysis-used.toml"
    written.append(_write_run_report(result, settings, outputfolder, written, when, fig_failures,
                                     cohort_outputs=cohort_outputs, manifest_name=manifest_name))
    return written


def write_planned(result, settings, plan, outputfolder: str | None = None,
                  when: datetime | None = None, progress=None) -> list[str]:
    """Write an ALREADY-COMPUTED ``result`` again, using ``write_batch``'s own, unchanged,
    golden-safe write logic — never recomputing the analysis (ticket A06 point 7: the Run
    screen's "Write results to another folder..." button, offered when the analysis
    succeeded but writing the first time failed, e.g. a folder that turned out to be
    read-only).

    ``plan`` supplies ``cohort_outputs`` (via ``plan.cohort_outputs``) rather than a
    separate parameter here, so a write can never disagree with the plan the user was shown
    for it — passing the flag independently was tried first and is exactly how a subset
    write's cohort-level outputs could get silently rebuilt from a plan that was never built
    for that case. ``outputfolder`` defaults to ``plan.outputfolder``; pass an explicit,
    DIFFERENT folder for "write elsewhere" so the plan that was computed against the
    original folder can still be reused (the write logic itself does not read the plan's
    ``outputfolder`` for anything but this default — every actual path is still built fresh
    from the ``outputfolder`` this call receives)."""
    target = outputfolder if outputfolder is not None else plan.outputfolder
    return write_batch(result, settings, target, when=when, progress=progress,
                       cohort_outputs=plan.cohort_outputs)


# --------------------------------------------------------------------------- #
# cohort summary (P8/P15)
# --------------------------------------------------------------------------- #
def _write_cohort_summary(result, settings, datadir, when) -> list[str]:
    agg = build_cohort_summary(result, settings)
    summary = agg.get("summary")
    if summary is None or len(summary) == 0:
        return []
    path = os.path.join(datadir, "Cohort summary.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        if agg.get("by_group") is not None:
            agg["by_group"].to_excel(writer, sheet_name="By group", index=False)
        _provenance_rows(settings, when).to_excel(writer, sheet_name="Provenance", index=False)
        _version_df().to_excel(writer, sheet_name="Version", index=False)
    return [path]


# --------------------------------------------------------------------------- #
# diagnostic figures (P11)
# --------------------------------------------------------------------------- #
def _write_figures(result, settings, outputfolder, progress=None, cohort_outputs: bool = True):
    """Returns (written_paths, failures). Failures — including matplotlib being
    absent — are surfaced in the run report, never silently dropped.

    The work goes to a separate process when the environment allows it: this runs on the
    GUI's worker thread, and matplotlib is not thread-safe while the Preview screen is using
    it on the GUI thread. See ``_figure_process`` — every failure mode falls back to doing it
    here, so the isolation can only ever be a no-op.

    ``progress`` is an optional ``callable(fname)`` fired once per file, used only on the
    in-process path (the child process can't call back across the boundary).

    ``cohort_outputs`` is forwarded to ``plots.write_figures`` — see ``write_batch``."""
    try:
        from respmech.core import plots  # noqa: F401 - import probe: is plotting available?
    except Exception as e:                       # pragma: no cover - plotting optional
        return [], [("figures", f"plotting unavailable: {e}")]
    from respmech.core.io import _figure_process
    notes = []
    written, failures = _figure_process.write_figures(
        result, settings, outputfolder, on_fallback=notes.append, progress=progress,
        cohort_outputs=cohort_outputs)
    # A fallback is not a failure — the figures are written either way — but it belongs in the
    # run report, because it is the difference between the isolated and the shared path.
    return written, failures + [("figures", n) for n in notes]


# --------------------------------------------------------------------------- #
# provenance files (P7)
# --------------------------------------------------------------------------- #
def _write_manifest(settings, outputfolder: str, *, cohort_outputs: bool = True,
                    when: datetime | None = None) -> str | None:
    """Drop the exact settings as a reloadable ``analysis-used.toml``, or ``None`` if
    nothing needed writing.

    A full run (``cohort_outputs=True``) always (over)writes ``analysis-used.toml`` —
    unchanged behaviour. A subset run must not silently replace the full run's provenance
    (A05): if a manifest already exists and is byte-identical to what this run's settings
    would produce, it is left alone (returns ``None`` — the existing file already describes
    this run accurately, so there is nothing new to report). If it exists and differs, the
    subset's settings are written under a distinct, timestamped name instead, so the full
    run's manifest survives untouched. If no manifest exists yet (nothing to protect — e.g.
    a "Process & write this file" into a folder that has never had a full run), it is
    written under the normal name, exactly as a full run would.

    A manifest that EXISTS but cannot be read back (permissions, a transient I/O error, or
    content that is not valid UTF-8 — e.g. hand-edited or from a crashed write) is treated
    the same as "exists and differs": protected under a timestamped name, never silently
    overwritten. Read failure must never be read as "nothing to protect" — that would
    reopen exactly the data loss this function exists to close."""
    from respmech.settingsio.toml_io import dumps_toml
    header = (f"# RespMech v{__version__} — settings used for this run.\n"
              f"# Reload with: File ▸ Load analysis (or point RespMech at this file).\n\n")
    content = header + dumps_toml(settings)
    full_path = os.path.join(outputfolder, "analysis-used.toml")
    if cohort_outputs:
        with open(full_path, "w", encoding="utf-8") as f:
            f.write(content)
        return full_path

    has_existing = os.path.isfile(full_path)
    existing = None
    if has_existing:
        try:
            with open(full_path, "r", encoding="utf-8") as f:
                existing = f.read()
        except (OSError, UnicodeDecodeError):
            existing = None                # unreadable — fall through to "protect it" below
    if existing is not None and existing == content:
        return None
    if not has_existing:
        with open(full_path, "w", encoding="utf-8") as f:
            f.write(content)
        return full_path
    ts = (when or datetime.now()).strftime("%Y%m%d-%H%M%S")
    partial_path = os.path.join(outputfolder, f"analysis-used (partial, {ts}).toml")
    with open(partial_path, "w", encoding="utf-8") as f:
        f.write(content)
    return partial_path


def _breath_counts(fr) -> tuple[int, int]:
    """(total, excluded) breath counts for a file result."""
    if not fr.breaths:
        return 0, 0
    total = len(fr.breaths)
    excluded = sum(1 for b in fr.breaths.values() if b.get("ignored"))
    return total, excluded


def _yn(flag) -> str:
    return "yes" if flag else "no"


def _write_run_report(result, settings, outputfolder: str,
                      written: list[str], when: datetime | None,
                      fig_failures=None, cohort_outputs: bool = True,
                      manifest_name: str = "analysis-used.toml") -> str:
    """A plain-text provenance log of what was read, kept, excluded and written.

    ``cohort_outputs=False`` (a subset/re-run, A05) writes this under a distinct,
    timestamped name instead of replacing the full run's ``run-report.txt`` — that file is
    the provenance record for the WHOLE study, and a subset run is not that. ``manifest_name``
    names whichever ``analysis-used*.toml`` this run actually produced (or left standing), so
    the "Settings snapshot" line always points at a file that really describes this run."""
    ts = (when or datetime.now()).strftime("%Y-%m-%d %H:%M:%S")
    ip, vol, samp = settings.input, settings.processing.volume, settings.processing.sampling
    seg, emg = settings.processing.segmentation, settings.processing.emg
    L: list[str] = []
    L.append(f"RespMech v{__version__} — run report")
    L.append(f"Generated: {ts}")
    if not cohort_outputs:
        L.append("")
        cohort_bits = []
        if settings.output.data.save_average:
            cohort_bits.append("Average breathdata.xlsx")
            cohort_bits.append("Cohort summary.xlsx")
        if settings.output.diagnostics.save_pv_individual:
            cohort_bits.append("the cohort Campbell figure")
        if cohort_bits:
            named = (cohort_bits[0] if len(cohort_bits) == 1
                    else ", ".join(cohort_bits[:-1]) + " and " + cohort_bits[-1])
            be, pron = ("is", "it") if len(cohort_bits) == 1 else ("are", "them")
            L.append(f"PARTIAL RUN — restricted to a subset of the study's files. "
                     f"{named} {be} UNCHANGED by this run; run the full batch to update {pron}.")
        else:
            L.append("PARTIAL RUN — restricted to a subset of the study's files.")
    L.append("")
    # A settings key this version reads differently from the one that wrote the analysis
    # changes the numbers below, so it is recorded at the TOP of every run it affected.
    for note in getattr(settings, "notices", ()) or ():
        L.append("SETTINGS UPGRADED SINCE THIS ANALYSIS WAS SAVED")
        L.append(f"  {note}")
        L.append("")
    L.append("INPUT")
    L.append(f"  Folder:   {ip.folder}")
    L.append(f"  Pattern:  {ip.files}")
    L.append(f"  Sampling: {ip.format.sampling_frequency} Hz")
    L.append("")

    ok, failed = result.ok_files, result.failed_files
    L.append(f"FILES ({len(ok)} processed, {len(failed)} failed)")
    for fname, fr in ok.items():
        total, excl = _breath_counts(fr)
        used = total - excl
        note = f" ({excl} excluded → {used} used)" if excl else ""
        L.append(f"  [ok]   {fname}   {total} breaths{note}")
    for fname, fr in failed.items():
        L.append(f"  [FAIL] {fname}   ERROR: {fr.error}")
    L.append("")

    L.append("PROCESSING")
    L.append(f"  Integrate flow → volume: {_yn(vol.integrate_from_flow)}")
    L.append(f"  Invert flow / volume:    {_yn(vol.inverse_flow)} / {_yn(vol.inverse_volume)}")
    L.append(f"  Drift correction:        {_yn(vol.correct_drift)}")
    if vol.correct_trend:
        # An "auto" analysis writes NO threshold key, so the run report is the only place
        # the rule that was actually applied is recorded.
        anchor = (f"absolute {vol.trend_peak_min_height:g} below the maximum, legacy"
                  if vol.trend_peak_min_height is not None
                  else f"breath depth ≥ {vol.trend_peak_min_prominence_frac:g} × volume range")
        L.append(f"  Trend correction:        Yes ({vol.trend_method}; {anchor}; troughs "
                 f"≥ {vol.trend_peak_min_distance_s:g} s apart)")
    else:
        L.append("  Trend correction:        No")
    L.append(f"  Resample:                {_yn(samp.resample)}"
             + (f" (→ {samp.resample_to_frequency} Hz)" if samp.resample else ""))
    L.append(f"  Breath separation:       by {seg.method}, buffer {seg.buffer}")
    L.append(f"  ECG removal:             {_yn(emg.remove_ecg)}")
    L.append(f"  EMG noise removal:       {_yn(emg.noise.enabled)}")
    L.append(f"  EMG normalisation:       {emg.normalization}")
    L.append("")

    # ECG / noise numeric diagnostics (audit #14): persist the R-peak counts, suppression,
    # chosen prop_decrease and per-channel fidelity/ΔSNR — previously only shown in-app.
    nr = getattr(result, "noise_report", None)
    ecg_auto = getattr(result, "ecg_auto_report", None)
    ecg_files = [(f, fr.ecg) for f, fr in ok.items() if getattr(fr, "ecg", None)]
    if nr or ecg_files or ecg_auto:
        L.append("DIAGNOSTICS")
        if ecg_auto:
            diag = ecg_auto.get("_diagnostics", {})
            bpm = diag.get("est_bpm")
            L.append(f"  ECG auto-detect: settings derived from {ecg_auto.get('reference_file')}, "
                     f"channel {ecg_auto.get('detect_channel')}, confidence {diag.get('confidence')}"
                     + (f" (~{bpm:.0f} bpm)" if bpm else "") + ". Check the per-file R-peak counts "
                     "below against this file count for any file the shared settings may not fit.")
        if ecg_files:
            L.append("  ECG removal (R-peaks captured / peak-window RMS suppression):")
            for f, d in ecg_files:
                supp = d.get("suppression", float("nan"))
                supp_s = f"{supp:.0%}" if supp == supp else "n/a"
                L.append(f"    {f}: {d.get('n_peaks', 0)} peaks (channel {d.get('detect_channel')}), "
                         f"suppression {supp_s}")
        if nr:
            L.append(f"  Noise reduction: prop_decrease {nr.get('prop_decrease')} "
                     f"(fidelity target {nr.get('fidelity_target')})")
            for ch in nr.get("channels", []):
                fid, dsnr = ch.get("fidelity"), ch.get("delta_snr_db")
                if fid is not None and dsnr is not None:
                    L.append(f"    channel {ch.get('channel')}: fidelity {fid:.3f}, ΔSNR {dsnr:+.1f} dB")
        L.append("")

    report_name = "run-report.txt"
    if not cohort_outputs:
        fts = (when or datetime.now()).strftime("%Y%m%d-%H%M%S")
        report_name = f"run-report (partial, {fts}).txt"

    L.append(f"OUTPUTS WRITTEN ({len(written) + 1} files)")
    for p in written:
        try:
            rel = os.path.relpath(p, outputfolder)
        except ValueError:               # Windows: p and outputfolder on different drives
            rel = os.path.basename(p)     # a cosmetic display path must never crash the report
        L.append(f"  {rel}")
    L.append(f"  {report_name}")
    L.append("")

    if fig_failures:
        L.append(f"FIGURES SKIPPED ({len(fig_failures)})")
        for name, err in fig_failures:
            L.append(f"  {name}: {err}")
        L.append("")

    L.append(f"Settings snapshot: {manifest_name} (reload to reproduce this run).")

    path = os.path.join(outputfolder, report_name)
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(L) + "\n")
    return path
