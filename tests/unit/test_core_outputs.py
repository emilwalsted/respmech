"""Core publishable-science outputs: units, cohort summary, EMG normalisation, and
diagnostic figures — all additive at the writer level so the golden result tables are
byte-identical (units P10/P21, cohort mean±SD/CV% P8, figures P11, normalisation P14,
grouping P15). Previously test_review_wave3.py."""
import os
from datetime import datetime
from types import SimpleNamespace

import numpy as np
import pandas as pd
import pytest

from _helpers import INPUT, requires_synth, synth_settings

pytestmark = requires_synth()


# --------------------------------------------------------------------------- #
# P10 / P21 — units
# --------------------------------------------------------------------------- #
def test_units_resolve_by_physical_quantity():
    from respmech.core import units
    assert units.unit_for("poes_mininsp") == "cmH₂O"
    assert units.unit_for("vt") == "L"
    assert units.unit_for("max_in_flow") == "L/s"       # 'flow' beats the volume rule
    assert units.unit_for("ti") == "s"
    assert units.unit_for("bf") == "min⁻¹"
    # WOB and PTP are scaled ×bcnt·vefactor in compute → per-minute quantities
    assert units.unit_for("wobtotal") == "J·min⁻¹"
    assert units.unit_for("ptp_oesinsp") == "cmH₂O·s·min⁻¹"   # rate, not the raw integral
    assert units.unit_for("int_oesinsp") == "cmH₂O·s"         # the un-scaled per-breath integral
    assert units.unit_for("sample_entropy_max") == "—"
    assert units.unit_for("file") == "" and units.unit_for("breath_no") == ""


# --------------------------------------------------------------------------- #
# P8 / P15 — cohort summary + grouping
# --------------------------------------------------------------------------- #
def test_cohort_summary_stats_and_grouping():
    from respmech.core.summary import build_cohort_summary, group_key
    assert group_key("P03_120W.csv") == "P03"
    assert group_key("S2 baseline.csv") == "S2"
    avg = pd.DataFrame({"file": ["S1_a.csv", "S1_b.csv", "S2_a.csv", "S2_b.csv"],
                        "wobtotal": [1.0, 2.0, 3.0, 4.0]})
    agg = build_cohort_summary(SimpleNamespace(average_table=avg),
                               SimpleNamespace(output=SimpleNamespace(group_regex=None)))
    row = agg["summary"].set_index("variable").loc["wobtotal"]
    assert row["n"] == 4 and row["mean"] == pytest.approx(2.5)
    assert row["sd"] == pytest.approx(np.std([1, 2, 3, 4], ddof=1))
    assert row["cv_pct"] == pytest.approx(100 * row["sd"] / row["mean"])
    bg = agg["by_group"].set_index(["group", "variable"]).loc[("S1", "wobtotal")]
    assert bg["mean"] == pytest.approx(1.5) and bg["n_files"] == 2


def test_grouping_regex_override():
    from respmech.core.summary import group_key
    s = SimpleNamespace(output=SimpleNamespace(group_regex=r"_(\d+)W"))
    assert group_key("subjX_120W.csv", s) == "120"
    assert group_key("nomatch.csv", s) == "(all)"       # falls back, never crashes


def test_grouping_optional_group_never_returns_none():
    """A regex whose capture group is optional must not return None (would break the
    sort in build_cohort_summary and abort the run after data was written)."""
    from respmech.core.summary import group_key, build_cohort_summary
    s = SimpleNamespace(output=SimpleNamespace(group_regex=r"P(\d+)?"))
    assert group_key("P_baseline.csv", s) == "P"        # group didn't participate → whole match
    assert group_key("P03.csv", s) == "03"
    avg = pd.DataFrame({"file": ["P_a.csv", "P03_b.csv"], "vt": [0.5, 0.6]})
    agg = build_cohort_summary(SimpleNamespace(average_table=avg), s)   # must not raise
    assert agg["by_group"] is not None


def test_cv_percent_is_nan_for_signed_mean():
    """CV% is undefined for a negative/near-zero mean; report NaN rather than a
    misleading negative percentage in a publishable summary."""
    from respmech.core.summary import build_cohort_summary
    avg = pd.DataFrame({"file": ["a.csv", "b.csv", "c.csv"],
                        "poes_mininsp": [-5.0, -7.0, -6.0], "vt": [0.5, 0.6, 0.55]})
    agg = build_cohort_summary(SimpleNamespace(average_table=avg),
                               SimpleNamespace(output=SimpleNamespace(group_regex=None)))
    stats = agg["summary"].set_index("variable")
    assert np.isnan(stats.loc["poes_mininsp", "cv_pct"])   # negative mean → no CV
    assert stats.loc["poes_mininsp", "mean"] == pytest.approx(-6.0)
    assert np.isfinite(stats.loc["vt", "cv_pct"])          # positive mean → real CV


def test_single_group_yields_no_comparison():
    from respmech.core.summary import build_cohort_summary
    avg = pd.DataFrame({"file": ["synth_a.csv", "synth_b.csv"], "vt": [0.5, 0.6]})
    agg = build_cohort_summary(SimpleNamespace(average_table=avg),
                               SimpleNamespace(output=SimpleNamespace(group_regex=None)))
    assert agg["by_group"] is None                      # one group -> nothing to compare


# --------------------------------------------------------------------------- #
# P14 — EMG normalisation (never mutates the raw table)
# --------------------------------------------------------------------------- #
def test_emg_normalization_modes():
    from respmech.core.summary import normalize_emg_table
    bt = pd.DataFrame({"breath_no": [1, 2, 3], "rms_max": [2.0, 4.0, 8.0], "vt": [1, 1, 1]})
    off = SimpleNamespace(processing=SimpleNamespace(emg=SimpleNamespace(normalization="none")))
    assert normalize_emg_table(bt, off) is None
    mx = SimpleNamespace(processing=SimpleNamespace(emg=SimpleNamespace(normalization="per_file_max")))
    nt = normalize_emg_table(bt, mx)
    assert list(nt["rms_max_pct"]) == pytest.approx([25.0, 50.0, 100.0])   # % of the peak breath
    assert "vt_pct" not in nt.columns                    # only RMS columns are normalised
    assert list(bt["rms_max"]) == [2.0, 4.0, 8.0]        # raw table untouched


# --------------------------------------------------------------------------- #
# P11 — diagnostic figures
# --------------------------------------------------------------------------- #
def test_figures_written_and_flag_driven(tmp_path):
    from respmech.core.pipeline import run_batch
    from respmech.core import plots
    s = synth_settings(tmp_path)
    s.output.diagnostics.save_pv_average = True
    s.output.diagnostics.save_pv_individual = False
    s.output.diagnostics.save_raw = False
    s.output.diagnostics.save_trimmed = False
    s.output.diagnostics.save_drift = False
    s.output.diagnostics.save_emg = False
    result = run_batch(s)
    written, failures = plots.write_figures(result, s, str(tmp_path))
    assert not failures
    names = [os.path.basename(p) for p in written]
    assert all(n.endswith("Campbell (average).pdf") for n in names)   # only the enabled figure, as PDF
    assert len(names) == len(result.ok_files)
    for p in written:
        assert os.path.getsize(p) > 0


def test_drift_figure_renders_from_volume_endpoints(tmp_path):
    """Regression: eelv/eilv are [volume, pressure] pairs, not scalars — the drift
    figure must render (it was silently failing and producing no file)."""
    from respmech.core.pipeline import run_batch
    from respmech.core import plots
    s = synth_settings(tmp_path)
    for flag in ("save_pv_average", "save_pv_individual", "save_raw", "save_trimmed", "save_emg"):
        setattr(s.output.diagnostics, flag, False)
    s.output.diagnostics.save_drift = True
    result = run_batch(s)
    written, failures = plots.write_figures(result, s, str(tmp_path))
    assert not failures                                  # no swallowed ValueError
    # save_drift now drives the staged volume-correction figure AND the endpoint check;
    # trend adjustment is off here, so its figure is skipped.
    assert written and all(p.endswith(".pdf") for p in written)
    assert any(p.endswith("volume endpoints.pdf") for p in written)   # the eelv/eilv endpoint figure
    assert any(p.endswith("volume correction.pdf") for p in written)


def test_emg_figures_wav_cohort_and_metrics(tmp_path):
    """Restored EMG overviews (one per conditioning stage), WAV export, the cohort
    average-Campbell, and the persisted ECG/noise diagnostics (audit #1/#10/#14/#15)."""
    from respmech.core.pipeline import run_batch
    from respmech.core import plots
    from respmech.core.io.writers import _write_run_report
    s = synth_settings(tmp_path, noise=True)             # ECG removal + shared-profile noise
    s.processing.emg.save_sound = True
    result = run_batch(s)
    written, failures = plots.write_figures(result, s, str(tmp_path))
    assert not failures
    names = [os.path.basename(p) for p in written]
    for stage in ("Raw EMG", "EMG (ECG removed)", "EMG (ECG removed + noise reduced)"):
        assert any(stage in n and n.endswith(".pdf") for n in names), f"missing EMG {stage} figure"
    assert any(n.endswith(".wav") for n in names)                    # WAV export
    assert any("All files" in n and n.endswith(".pdf") for n in names)   # cohort Campbell
    rp = _write_run_report(result, s, str(tmp_path), written, None)
    txt = open(rp, encoding="utf-8").read()
    assert "DIAGNOSTICS" in txt and "prop_decrease" in txt and "R-peaks" in txt


def test_volume_correction_figure_omits_drift_panel_when_drift_off(tmp_path, monkeypatch):
    """The staged volume-correction figure must not show a 'drift-corrected' panel that is
    byte-identical to 'zeroed' when drift correction is off (review nit)."""
    from respmech.core.pipeline import run_batch
    from respmech.core import plots
    s = synth_settings(tmp_path)
    s.processing.volume.correct_drift = False
    for flag in ("save_pv_average", "save_pv_individual", "save_raw", "save_trimmed", "save_emg"):
        setattr(s.output.diagnostics, flag, False)
    s.output.diagnostics.save_drift = True
    result = run_batch(s)
    fr = next(iter(result.ok_files.values()))
    rows = []
    monkeypatch.setattr(plots, "_save", lambda fig, path: (rows.append([ax.get_ylabel() for ax in fig.axes]), path)[1])
    plots._volume_correction(fr, "f", "/tmp/vc.pdf")
    labels = rows[0]
    assert "Linear drift-corrected (L)" not in labels          # skipped when drift is off
    assert "Zeroed volume (L)" in labels


def test_save_emg_flag_gates_emg_figures(tmp_path):
    from respmech.core.pipeline import run_batch
    from respmech.core import plots
    s = synth_settings(tmp_path, remove_ecg=True)
    for flag in ("save_pv_average", "save_pv_individual", "save_raw", "save_trimmed", "save_drift"):
        setattr(s.output.diagnostics, flag, False)
    s.output.diagnostics.save_emg = True
    written, _ = plots.write_figures(run_batch(s), s, str(tmp_path))
    assert written and all("EMG" in os.path.basename(p) for p in written)
    s.output.diagnostics.save_emg = False
    written2, _ = plots.write_figures(run_batch(s), s, str(tmp_path))
    assert written2 == []                                            # nothing when all off


# --------------------------------------------------------------------------- #
# writer integration — additive, golden-safe
# --------------------------------------------------------------------------- #
def test_write_batch_adds_units_provenance_summary_without_touching_data(tmp_path):
    from respmech.core.pipeline import run_batch
    from respmech.core.io.writers import write_batch
    s = synth_settings(tmp_path)
    result = run_batch(s)
    raw_cols = list(next(iter(result.ok_files.values())).breaths_table.columns)
    write_batch(result, s, str(tmp_path), when=datetime(2026, 7, 11))

    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(tmp_path, "data", "synth_case_A.csv.breathdata.xlsx"))
    assert {"Data", "Units", "Provenance", "Version"} <= set(wb.sheetnames)
    assert "EMG normalised" in wb.sheetnames                # default per-file max, EMG present
    # the Data sheet still holds exactly the pinned columns — nothing added/renamed there
    data_cols = [c.value for c in next(wb["Data"].iter_rows(max_row=1))]
    assert data_cols == raw_cols
    assert os.path.isfile(os.path.join(tmp_path, "data", "Cohort summary.xlsx"))


# ---------------------------------------------------------------------------
# Run report + reloadable manifest (P7); sample data (P23) — waves 2, 6
# ---------------------------------------------------------------------------
def test_run_report_accounts_for_excluded_and_failed(tmp_path):
    """The run report must show per-file breath accounting (excluded → used) and any
    failures, without needing a full batch — driven by lightweight fakes."""
    from types import SimpleNamespace
    from respmech.core.io.writers import _write_run_report

    ok = SimpleNamespace(breaths={1: {"ignored": False}, 2: {"ignored": True},
                                  3: {"ignored": False}}, error=None)
    bad = SimpleNamespace(breaths=None, error="boom while loading")
    result = SimpleNamespace(
        ok_files={"good.csv": ok}, failed_files={"bad.csv": bad})
    s = synth_settings(tmp_path)
    path = _write_run_report(result, s, str(tmp_path), ["data/x.xlsx"], datetime(2026, 7, 11))
    report = open(path, encoding="utf-8").read()   # the report is UTF-8 (em-dash, → arrow)
    assert "1 processed, 1 failed" in report
    assert "3 breaths (1 excluded → 2 used)" in report
    assert "[FAIL] bad.csv   ERROR: boom while loading" in report


def test_write_batch_emits_reloadable_manifest(tmp_path):
    from respmech.core.pipeline import run_batch
    from respmech.core.io.writers import write_batch
    from respmech.settingsio.toml_io import load_toml
    s = synth_settings(tmp_path)
    s.output.data.save_processed = True
    result = run_batch(s)
    written = write_batch(result, s, str(tmp_path), when=datetime(2026, 7, 11, 14, 0, 0))
    manifest = os.path.join(str(tmp_path), "analysis-used.toml")
    report = os.path.join(str(tmp_path), "run-report.txt")
    assert os.path.isfile(manifest) and os.path.isfile(report)
    assert manifest in written and report in written
    # the manifest is a real, reloadable settings file (round-trips the sampling rate)
    reloaded = load_toml(manifest)
    assert reloaded.input.format.sampling_frequency == 1000
    assert "2 processed, 0 failed" in open(report, encoding="utf-8").read()


def test_sample_recording_is_analysable(tmp_path):
    from respmech.core.sample import write_sample_recording
    from respmech.core.pipeline import run_batch
    from respmech.core.settings import Settings
    desc = write_sample_recording(str(tmp_path / "input"))
    assert os.path.isfile(desc["path"])
    s = Settings()
    s.input.folder = desc["folder"]; s.input.files = desc["filename"]
    s.input.format.sampling_frequency = desc["sampling_frequency"]
    m, ch = desc["mapping"], s.input.channels
    ch.flow, ch.volume, ch.poes, ch.pgas, ch.pdi = m["flow"], m["volume"], m["poes"], m["pgas"], m["pdi"]
    ch.emg = m["emg"]; s.processing.segmentation.buffer = 200
    result = run_batch(s)
    fr = next(iter(result.ok_files.values()))
    assert len([b for b in fr.breaths.values() if not b["ignored"]]) >= 4   # real breaths detected


def test_output_pdfs_are_always_light(monkeypatch):
    """An output figure is a document that leaves the app, not a view of it. The GUI's dark
    theme installs its colours into GLOBAL matplotlib rcParams, which core.plots would
    otherwise inherit at Figure()/savefig() time — so a dark UI must not bleed into a
    written PDF, and rendering must not leave the GUI's own canvases light afterwards."""
    import matplotlib as mpl
    from respmech.core import plot_style
    from respmech.ui import theme
    # install the DARK theme's rcParams the way a running GUI does
    monkeypatch.setattr(theme, "_IS_DARK", True)
    theme.style_matplotlib()
    dark_ground = mpl.rcParams["axes.facecolor"]
    assert dark_ground != "#FFFFFF"                              # the leak source is armed

    with plot_style.light_rc_context():                           # what write_figures wraps in
        assert mpl.rcParams["axes.facecolor"] == "#FFFFFF"
        assert mpl.rcParams["figure.facecolor"] == "#FFFFFF"
        assert mpl.rcParams["text.color"] == "#33404D"
    assert mpl.rcParams["axes.facecolor"] == dark_ground          # GUI canvases stay dark

    # ...and write_figures really RUNS inside that context — this is the wiring the whole
    # fix hangs on, so pin it: spy on the impl and read the live rcParams from within.
    from respmech.core import plots
    seen = {}
    monkeypatch.setattr(plots, "_write_figures_impl",
                        lambda r, s, o, progress=None: (seen.update(face=mpl.rcParams["axes.facecolor"]),
                                                        ([], []))[1])
    assert plots.write_figures(None, None, "") == ([], [])
    assert seen["face"] == "#FFFFFF"                              # light INSIDE the writer
    assert mpl.rcParams["axes.facecolor"] == dark_ground          # ...dark restored after

    # The light table must cover EVERY key the theme re-colours, or that key leaks through.
    dark_rc = {k: mpl.rcParams[k] for k in plot_style.light_rc() if k != "axes.prop_cycle"}
    monkeypatch.setattr(theme, "_IS_DARK", False)
    theme.style_matplotlib()
    light = plot_style.light_rc()
    for key, dark_val in dark_rc.items():
        if dark_val != mpl.rcParams[key]:                         # a key the theme re-colours
            assert light[key] == mpl.rcParams[key], (
                f"{key}: core's light table has drifted from the light theme")
