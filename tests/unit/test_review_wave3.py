"""Wave 3 of the full-app review: publishable science.

All of it is additive at the writer level so the golden-pinned result tables never
change: P8 cohort mean ± SD / n / CV%; P10/P21 unit-labelled headers + a Units and a
Provenance sheet; P11 diagnostic figures; P14 EMG amplitude normalisation (a separate
normalised sheet, default per-file max); P15 subject/condition grouping + comparison.
"""
import os
from datetime import datetime
from types import SimpleNamespace

import numpy as np
import pandas as pd
import pytest

ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
INPUT = os.path.join(ROOT, "tests", "golden", "input")
pytestmark = pytest.mark.skipif(
    not os.path.exists(os.path.join(INPUT, "synth_case_A.csv")), reason="synthetic input absent")


def _legacy(out, **overrides):
    from respmech.settingsio.migrate import migrate_dict
    legacy = {"input": {"inputfolder": INPUT, "files": "synth_case_*.csv",
                        "format": {"samplingfrequency": 1000},
                        "data": {"column_poes": 7, "column_pgas": 8, "column_pdi": 9,
                                 "column_volume": 6, "column_flow": 5, "columns_emg": [2, 3, 4],
                                 "columns_entropy": [10, 11, 12]}},
              "processing": {"mechanics": {"breathseparationbuffer": 200, "separateby": "flow",
                                           "avgresamplingobs": 300},
                             "emg": {"remove_ecg": False, "remove_noise": False}},
              "output": {"outputfolder": str(out), "data": {}}}
    s, _ = migrate_dict(legacy)
    s.input.folder = INPUT
    s.output.folder = str(out)
    for k, v in overrides.items():
        setattr(s, k, v)
    return s


# --------------------------------------------------------------------------- #
# P10 / P21 — units
# --------------------------------------------------------------------------- #
def test_units_resolve_by_physical_quantity():
    from respmech.core import units
    assert units.unit_for("poes_mininsp") == "cmH₂O"
    assert units.unit_for("vt") == "L"
    assert units.unit_for("max_in_flow") == "L/s"       # 'flow' beats the volume rule
    assert units.unit_for("ti") == "s"
    assert units.unit_for("bf") == "min⁻¹"
    # WOB and PTP are scaled ×bcnt·vefactor in compute → per-minute quantities
    assert units.unit_for("wobtotal") == "J·min⁻¹"
    assert units.unit_for("ptp_oesinsp") == "cmH₂O·s·min⁻¹"   # rate, not the raw integral
    assert units.unit_for("int_oesinsp") == "cmH₂O·s"         # the un-scaled per-breath integral
    assert units.unit_for("sample_entropy_max") == "—"
    assert units.unit_for("file") == "" and units.unit_for("breath_no") == ""


# --------------------------------------------------------------------------- #
# P8 / P15 — cohort summary + grouping
# --------------------------------------------------------------------------- #
def test_cohort_summary_stats_and_grouping():
    from respmech.core.summary import build_cohort_summary, group_key
    assert group_key("P03_120W.csv") == "P03"
    assert group_key("S2 baseline.csv") == "S2"
    avg = pd.DataFrame({"file": ["S1_a.csv", "S1_b.csv", "S2_a.csv", "S2_b.csv"],
                        "wobtotal": [1.0, 2.0, 3.0, 4.0]})
    agg = build_cohort_summary(SimpleNamespace(average_table=avg),
                               SimpleNamespace(output=SimpleNamespace(group_regex=None)))
    row = agg["summary"].set_index("variable").loc["wobtotal"]
    assert row["n"] == 4 and row["mean"] == pytest.approx(2.5)
    assert row["sd"] == pytest.approx(np.std([1, 2, 3, 4], ddof=1))
    assert row["cv_pct"] == pytest.approx(100 * row["sd"] / row["mean"])
    bg = agg["by_group"].set_index(["group", "variable"]).loc[("S1", "wobtotal")]
    assert bg["mean"] == pytest.approx(1.5) and bg["n_files"] == 2


def test_grouping_regex_override():
    from respmech.core.summary import group_key
    s = SimpleNamespace(output=SimpleNamespace(group_regex=r"_(\d+)W"))
    assert group_key("subjX_120W.csv", s) == "120"
    assert group_key("nomatch.csv", s) == "(all)"       # falls back, never crashes


def test_grouping_optional_group_never_returns_none():
    """A regex whose capture group is optional must not return None (would break the
    sort in build_cohort_summary and abort the run after data was written)."""
    from respmech.core.summary import group_key, build_cohort_summary
    s = SimpleNamespace(output=SimpleNamespace(group_regex=r"P(\d+)?"))
    assert group_key("P_baseline.csv", s) == "P"        # group didn't participate → whole match
    assert group_key("P03.csv", s) == "03"
    avg = pd.DataFrame({"file": ["P_a.csv", "P03_b.csv"], "vt": [0.5, 0.6]})
    agg = build_cohort_summary(SimpleNamespace(average_table=avg), s)   # must not raise
    assert agg["by_group"] is not None


def test_cv_percent_is_nan_for_signed_mean():
    """CV% is undefined for a negative/near-zero mean; report NaN rather than a
    misleading negative percentage in a publishable summary."""
    from respmech.core.summary import build_cohort_summary
    avg = pd.DataFrame({"file": ["a.csv", "b.csv", "c.csv"],
                        "poes_mininsp": [-5.0, -7.0, -6.0], "vt": [0.5, 0.6, 0.55]})
    agg = build_cohort_summary(SimpleNamespace(average_table=avg),
                               SimpleNamespace(output=SimpleNamespace(group_regex=None)))
    stats = agg["summary"].set_index("variable")
    assert np.isnan(stats.loc["poes_mininsp", "cv_pct"])   # negative mean → no CV
    assert stats.loc["poes_mininsp", "mean"] == pytest.approx(-6.0)
    assert np.isfinite(stats.loc["vt", "cv_pct"])          # positive mean → real CV


def test_single_group_yields_no_comparison():
    from respmech.core.summary import build_cohort_summary
    avg = pd.DataFrame({"file": ["synth_a.csv", "synth_b.csv"], "vt": [0.5, 0.6]})
    agg = build_cohort_summary(SimpleNamespace(average_table=avg),
                               SimpleNamespace(output=SimpleNamespace(group_regex=None)))
    assert agg["by_group"] is None                      # one group -> nothing to compare


# --------------------------------------------------------------------------- #
# P14 — EMG normalisation (never mutates the raw table)
# --------------------------------------------------------------------------- #
def test_emg_normalization_modes():
    from respmech.core.summary import normalize_emg_table
    bt = pd.DataFrame({"breath_no": [1, 2, 3], "rms_max": [2.0, 4.0, 8.0], "vt": [1, 1, 1]})
    off = SimpleNamespace(processing=SimpleNamespace(emg=SimpleNamespace(normalization="none")))
    assert normalize_emg_table(bt, off) is None
    mx = SimpleNamespace(processing=SimpleNamespace(emg=SimpleNamespace(normalization="per_file_max")))
    nt = normalize_emg_table(bt, mx)
    assert list(nt["rms_max_pct"]) == pytest.approx([25.0, 50.0, 100.0])   # % of the peak breath
    assert "vt_pct" not in nt.columns                    # only RMS columns are normalised
    assert list(bt["rms_max"]) == [2.0, 4.0, 8.0]        # raw table untouched


# --------------------------------------------------------------------------- #
# P11 — diagnostic figures
# --------------------------------------------------------------------------- #
def test_figures_written_and_flag_driven(tmp_path):
    from respmech.core.pipeline import run_batch
    from respmech.core import plots
    s = _legacy(tmp_path)
    s.output.diagnostics.save_pv_average = True
    s.output.diagnostics.save_pv_individual = False
    s.output.diagnostics.save_raw = False
    s.output.diagnostics.save_trimmed = False
    s.output.diagnostics.save_drift = False
    result = run_batch(s)
    written, failures = plots.write_figures(result, s, str(tmp_path))
    assert not failures
    names = [os.path.basename(p) for p in written]
    assert all(n.endswith("PV loops (average).png") for n in names)   # only the enabled figure
    assert len(names) == len(result.ok_files)
    for p in written:
        assert os.path.getsize(p) > 0


def test_drift_figure_renders_from_volume_endpoints(tmp_path):
    """Regression: eelv/eilv are [volume, pressure] pairs, not scalars — the drift
    figure must render (it was silently failing and producing no file)."""
    from respmech.core.pipeline import run_batch
    from respmech.core import plots
    s = _legacy(tmp_path)
    for flag in ("save_pv_average", "save_pv_individual", "save_raw", "save_trimmed"):
        setattr(s.output.diagnostics, flag, False)
    s.output.diagnostics.save_drift = True
    result = run_batch(s)
    written, failures = plots.write_figures(result, s, str(tmp_path))
    assert not failures                                  # no swallowed ValueError
    assert written and all(p.endswith("drift.png") for p in written)
    assert len(written) == len(result.ok_files)


# --------------------------------------------------------------------------- #
# writer integration — additive, golden-safe
# --------------------------------------------------------------------------- #
def test_write_batch_adds_units_provenance_summary_without_touching_data(tmp_path):
    from respmech.core.pipeline import run_batch
    from respmech.core.io.writers import write_batch
    s = _legacy(tmp_path)
    result = run_batch(s)
    raw_cols = list(next(iter(result.ok_files.values())).breaths_table.columns)
    write_batch(result, s, str(tmp_path), when=datetime(2026, 7, 11))

    import openpyxl
    wb = openpyxl.load_workbook(os.path.join(tmp_path, "data", "synth_case_A.csv.breathdata.xlsx"))
    assert {"Data", "Units", "Provenance", "Version"} <= set(wb.sheetnames)
    assert "EMG normalised" in wb.sheetnames                # default per-file max, EMG present
    # the Data sheet still holds exactly the pinned columns — nothing added/renamed there
    data_cols = [c.value for c in next(wb["Data"].iter_rows(max_row=1))]
    assert data_cols == raw_cols
    assert os.path.isfile(os.path.join(tmp_path, "data", "Cohort summary.xlsx"))
